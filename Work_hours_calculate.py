from __future__ import annotations
from datetime import datetime, time, timedelta, date
import calendar
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from dataclasses import dataclass
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
# ---------------------------------------------
# Core work log calculation
# ---------------------------------------------
@dataclass(frozen=True)
class WorkEntry:
    start: datetime
    end: datetime

    @property
    def duration(self) -> timedelta:
        return self.end - self.start

    @property
    def hours(self) -> float:
        return self.duration.total_seconds() / 3600.0

    @property
    def is_newday(self) -> bool:
        return self.end.date() != self.start.date()
def _add_page_number(c: canvas.Canvas, doc):
    c.saveState()
    c.setFont("Helvetica", 9)
    page_text = f"Page {doc.page}"
    c.drawRightString(A4[0] - 15 * mm, 12 * mm, page_text)
    c.restoreState()

class work_log:
    def __init__(self):
        self.entries: list[tuple[datetime, datetime]] = []
      
    def add_a_day_entry(self, date, start_time, end_time, is_newday:bool=False):
        start_dt = datetime.combine(date, start_time)
        end_dt = datetime.combine(date, end_time)
        # If user says it's next day OR end is earlier than start, roll end forward by 1 day
        if is_newday or end_dt < start_dt:
            end_dt += timedelta(days=1)
        self.entries.append((start_dt, end_dt))
        return self.entries

    def calculate_total_hours(self):
        total = timedelta(0)
        for start_dt, end_dt in self.entries:
            total += (end_dt - start_dt)
        return total.total_seconds() / 3600.0
    
    
    def to_printable_text(self) -> str:
        lines = []
        lines.append(f"{'Date':10}  {'In':5}  {'Out':5}  {'NewDay':6}  {'Hours':>7}")
        lines.append("-" * 42)

        for start_dt, end_dt in self._iter_as_datetimes():
            hours = (end_dt - start_dt).total_seconds() / 3600.0
            is_newday = end_dt.date() != start_dt.date()
            lines.append(
                f"{start_dt.strftime('%Y%m%d'):10}  "
                f"{start_dt.strftime('%H:%M'):5}  "
                f"{end_dt.strftime('%H:%M'):5}  "
                f"{('Yes' if is_newday else 'No'):6}  "
                f"{hours:7.2f}"
            )

        lines.append("-" * 42)
        lines.append(f"TOTAL HOURS: {self.calculate_total_hours():.2f}")
        return "\n".join(lines)

    def export_to_excel(self, filepath: str) -> None:
        # pip install openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Work Log"

        headers = ["Date (yyyymmdd)", "Check-in", "Check-out", "New day?", "Start", "End", "Hours"]
        ws.append(headers)

        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for start_dt, end_dt in self._iter_as_datetimes():
            hours = (end_dt - start_dt).total_seconds() / 3600.0
            is_newday = end_dt.date() != start_dt.date()

            ws.append([
                start_dt.strftime("%Y%m%d"),
                start_dt.strftime("%H:%M"),
                end_dt.strftime("%H:%M"),
                "Yes" if is_newday else "No",
                start_dt,
                end_dt,
                round(hours, 2),
            ])

        # Column widths + formats
        widths = [16, 12, 12, 10, 20, 20, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=5).number_format = "yyyy-mm-dd hh:mm"
            ws.cell(row=r, column=6).number_format = "yyyy-mm-dd hh:mm"
            ws.cell(row=r, column=7).number_format = "0.00"

        # Total row
        last = ws.max_row + 1
        ws.cell(row=last, column=6, value="TOTAL").font = header_font
        ws.cell(row=last, column=7, value=round(self.calculate_total_hours(), 2)).font = header_font

        wb.save(filepath)

    def export_to_pdf(self, filepath: str, title: str = "Work Hours Report") -> None:
        styles = getSampleStyleSheet()

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=18 * mm,
        )

        elements = []
        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Spacer(1, 6 * mm))

        # Build table data
        data = [["Date (yyyymmdd)", "Check-in", "Check-out", "New day?", "Start", "End", "Hours"]]

        for start_dt, end_dt in self._iter_as_datetimes():
            hours = (end_dt - start_dt).total_seconds() / 3600.0
            is_newday = end_dt.date() != start_dt.date()

            data.append([
                start_dt.strftime("%Y%m%d"),
                start_dt.strftime("%H:%M"),
                end_dt.strftime("%H:%M"),
                "Yes" if is_newday else "No",
                start_dt.strftime("%Y-%m-%d %H:%M"),
                end_dt.strftime("%Y-%m-%d %H:%M"),
                f"{hours:.2f}",
            ])

        total_hours = self.calculate_total_hours()
        data.append(["", "", "", "", "", "TOTAL", f"{total_hours:.2f}"])

        # Table styling
        col_widths = [28*mm, 18*mm, 18*mm, 18*mm, 38*mm, 38*mm, 16*mm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)

        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),

            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.lightgrey]),

            # Total row emphasis
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.beige),
            ("ALIGN", (5, -1), (5, -1), "RIGHT"),
            ("ALIGN", (6, -1), (6, -1), "RIGHT"),
        ]))

        elements.append(tbl)

        doc.build(
            elements,
            onFirstPage=_add_page_number,
            onLaterPages=_add_page_number
        )
    def _iter_as_datetimes(self):
        """
        Yield (start_dt, end_dt) for each entry, regardless of whether it is:
        - WorkEntry with .start/.end
        - tuple(start_dt, end_dt)
        - tuple(day, start_time, end_time)  [end_time may be time]
        """
        for entry in self.entries:
            # Case A: WorkEntry-like object
            if hasattr(entry, "start") and hasattr(entry, "end"):
                yield entry.start, entry.end
                continue

            # Case B: tuple forms
            if isinstance(entry, tuple):
                if len(entry) == 2:
                    start_dt, end_dt = entry
                    yield start_dt, end_dt
                    continue

                if len(entry) == 3:
                    day, start_t, end_t = entry
                    start_dt = datetime.combine(day, start_t)

                    # end_t might be time or datetime
                    if isinstance(end_t, datetime):
                        end_dt = end_t
                    else:
                        end_dt = datetime.combine(day, end_t)

                    # auto-handle cross-midnight
                    if end_dt < start_dt:
                        end_dt += timedelta(days=1)

                    yield start_dt, end_dt
                    continue

            raise TypeError(f"Unsupported entry format: {entry!r}")
# -----------------------------
# Helpers (parsing/validation)
# -----------------------------
def parse_yyyymmdd(s: str) -> date:
    s = s.strip()
    # strict format: yyyymmdd
    return datetime.strptime(s, "%Y%m%d").date()

def parse_hhmm(s: str) -> time:
    s = s.strip()
    # strict format: HH:MM (24h)
    return datetime.strptime(s, "%H:%M").time()


# -----------------------------
# UI
# -----------------------------
class WorkLogApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Work Hours Calculator")
        self.geometry("820x500")

        self.work_log = work_log()
        self.rows: list[dict] = []

        # Top controls
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Button(top, text="Add Row", command=lambda: self.add_row(copy_last=False)).pack(side="left")

        ttk.Button(top, text="Add Row (Copy Last)",
                command=lambda: self.add_row(copy_last=True)).pack(side="left", padx=6)

        self.auto_copy_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text="Auto-copy last row", variable=self.auto_copy_var).pack(side="left", padx=6)

        ttk.Button(top, text="Calculate Total Hours", command=self.calculate_total).pack(side="left", padx=8)
        ttk.Button(top, text="Clear All", command=self.clear_all).pack(side="left")

        self.total_var = tk.StringVar(value="Total hours: 0.00")
        ttk.Label(top, textvariable=self.total_var, font=("Segoe UI", 11, "bold")).pack(side="right")
        ttk.Button(top, text="Print", command=self.print_report).pack(side="left", padx=6)
        ttk.Button(top, text="Export Excel", command=self.export_excel).pack(side="left", padx=6)
        tk.Button(top, text="Export PDF", command=self.export_pdf).pack(side="left", padx=6)

        # Table header
        header = ttk.Frame(self, padding=(10, 0, 10, 0))
        header.pack(fill="x")

        ttk.Label(header, text="Date (yyyymmdd)", width=16).grid(row=0, column=0, sticky="w")
        ttk.Label(header, text="Check-in", width=16).grid(row=0, column=1, sticky="w")
        ttk.Label(header, text="Check-out", width=16).grid(row=0, column=2, sticky="w")
        ttk.Label(header, text="New day", width=10).grid(row=0, column=3, sticky="w")
        ttk.Label(header, text="Status", width=40).grid(row=0, column=4, sticky="w")

        # Scrollable rows area
        container = ttk.Frame(self, padding=10)
        container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(container, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.scrollable = ttk.Frame(self.canvas)

        self.scrollable.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # Add initial row
        self.add_row()

    def add_row(self, copy_last: bool = False):
    # If auto-copy toggle is enabled, treat new rows as copy_last
        if hasattr(self, "auto_copy_var") and self.auto_copy_var.get():
            copy_last = True

        last = self.rows[-1] if (copy_last and self.rows) else None

        row_frame = ttk.Frame(self.scrollable)
        row_frame.pack(fill="x", pady=3)

        date_entry = ttk.Entry(row_frame, width=16)
        in_entry = ttk.Entry(row_frame, width=16)
        out_entry = ttk.Entry(row_frame, width=16)
        newday_var = tk.BooleanVar(value=False)
        newday_chk = ttk.Checkbutton(row_frame, variable=newday_var)

        status_var = tk.StringVar(value="")
        status_lbl = ttk.Label(row_frame, textvariable=status_var, width=40)

        date_entry.grid(row=0, column=0, padx=(0, 8), sticky="w")
        in_entry.grid(row=0, column=1, padx=(0, 8), sticky="w")
        out_entry.grid(row=0, column=2, padx=(0, 8), sticky="w")
        newday_chk.grid(row=0, column=3, padx=(0, 8), sticky="w")
        status_lbl.grid(row=0, column=4, sticky="w")

        rm_btn = ttk.Button(row_frame, text="Remove", command=lambda: self.remove_row(row_frame))
        rm_btn.grid(row=0, column=5, padx=(10, 0))

        # ---- Copy values from last row (if requested) ----
        if last is not None:
            date_entry.insert(0, last["date_entry"].get().strip())
            in_entry.insert(0, last["in_entry"].get().strip())
            out_entry.insert(0, last["out_entry"].get().strip())
            newday_var.set(bool(last["newday_var"].get()))

            # Optional UX: auto-focus date or check-in; choose check-in for speed
            in_entry.focus_set()
            in_entry.select_range(0, tk.END)
        else:
            date_entry.focus_set()

        self.rows.append({
            "frame": row_frame,
            "date_entry": date_entry,
            "in_entry": in_entry,
            "out_entry": out_entry,
            "newday_var": newday_var,
            "status_var": status_var,
        })

        self.after(50, lambda: self.canvas.yview_moveto(1.0))

    def remove_row(self, frame: ttk.Frame):
        # Remove from list and destroy widgets
        idx = None
        for i, r in enumerate(self.rows):
            if r["frame"] is frame:
                idx = i
                break
        if idx is not None:
            self.rows[idx]["frame"].destroy()
            self.rows.pop(idx)

    def clear_all(self):
        for r in self.rows:
            r["frame"].destroy()
        self.rows.clear()
        self.work_log = work_log()
        self.total_var.set("Total hours: 0.00")
        self.add_row()

    def calculate_total(self):
        self.work_log = work_log()
        any_errors = False

        for r in self.rows:
            r["status_var"].set("")  # clear status

            d_str = r["date_entry"].get().strip()
            in_str = r["in_entry"].get().strip()
            out_str = r["out_entry"].get().strip()
            newday = bool(r["newday_var"].get())

            # Skip completely empty rows (optional behavior)
            if not d_str and not in_str and not out_str:
                continue

            try:
                day = parse_yyyymmdd(d_str)
                t_in = parse_hhmm(in_str)
                t_out = parse_hhmm(out_str)

                self.work_log.add_a_day_entry(day, t_in, t_out, is_newday=newday)
                r["status_var"].set("OK")

            except Exception as e:
                any_errors = True
                r["status_var"].set(f"Error: {e}")

        total_hours = self.work_log.calculate_total_hours()
        self.total_var.set(f"Total hours: {total_hours:.2f}")

        if any_errors:
            messagebox.showwarning("Validation", "Some rows have errors. Please fix them and calculate again.")
    def print_report(self):
    # Ensure the log is up to date
        self.calculate_total()
        print(self.work_log.to_printable_text())

    def export_excel(self):
        # Ensure the log is up to date
        self.calculate_total()

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save work log as Excel"
        )
        if not path:
            return

        try:
            self.work_log.export_to_excel(path)
            messagebox.showinfo("Export", f"Saved Excel file:\n{path}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    def export_pdf(self):
        # Rebuild the WorkLog from current UI rows (and validate)
        self.calculate_total()

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Save report as PDF"
        )
        if not path:
            return

        try:
            self.work_log.export_to_pdf(path, title="Work Hours Report")
            messagebox.showinfo("Export PDF", f"Saved PDF file:\n{path}")
        except Exception as e:
            messagebox.showerror("Export PDF failed", str(e))
if __name__ == "__main__":
    app = WorkLogApp()
    app.mainloop()